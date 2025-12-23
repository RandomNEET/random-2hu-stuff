import sqlite3

import openpyxl
from openpyxl.styles import Color, Font


def export_authors_to_xlsx(db_path, output_xlsx):
    base_url = "https://random-2hu-stuff.randomneet.me/author/"

    try:
        # 1. 连接数据库
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        query = """
        SELECT 
            COALESCE(NULLIF(yt_name, ''), NULLIF(nico_name, ''), NULLIF(twitter_name, '')) AS author,
            id,
            comment
        FROM authors
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        # 2. 创建 XLSX 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Authors"

        # 写入表头
        ws.append(["作者", "链接", "备注"])

        # 3. 写入数据并设置超链接
        for row_num, (name, author_id, comment) in enumerate(rows, start=2):
            full_url = f"{base_url}{author_id}"

            # 写入作者名
            ws.cell(row=row_num, column=1, value=name)

            # 写入链接列并添加超链接属性
            cell = ws.cell(row=row_num, column=2, value=full_url)
            cell.hyperlink = full_url
            cell.font = Font(color="0000FF", underline="single")  # 设置为蓝色下划线样式

            # 写入备注
            ws.cell(row=row_num, column=3, value=comment)

        # 4. 保存文件
        wb.save(output_xlsx)
        print(f"导出成功！文件已保存为: {output_xlsx}")

    except Exception as e:
        print(f"发生错误: {e}")
    finally:
        if conn:
            conn.close()


# --- 执行 ---
export_authors_to_xlsx("../backend/random-2hu-stuff.db", "authors-comments.xlsx")
